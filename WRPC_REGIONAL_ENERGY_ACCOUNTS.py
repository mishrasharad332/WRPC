from openpyxl import Workbook, load_workbook
from datetime import datetime
import streamlit as st
import pandas as pd
import requests
import pdfplumber  # For PDF processing
import os
import io

def create_file(df, sheet_name):
    filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    # Check file existence
    if not os.path.exists(filename):
        wb = Workbook()
        wb.save(filename)
    else:
        wb = load_workbook(filename)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws2 = wb[sheet_name]

    # Write specific column names as headers
    headers = ["RE Generator", "Schedule (MU)", "Actual (MU)", "Deviation (MU)", "Year", "PDF URL"]
    ws2.append(headers)

    # Append data to the worksheet
    for index, row in df.iterrows():  # Iterate over DataFrame rows
        # Split the 'PDF URL' column into two separate columns for hyperlink function
        row["PDF URL"] = row["PDF URL"].split(", ")[1].strip(')')
        row_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(row_list)  # Append the row to the worksheet

    # Save the workbook
    wb.save(filename)

# Function to search for text in PDF and extract the row
def search_text_in_pdf(title, url, search_text):
    # Download the PDF file
    response = requests.get(url)
    pdf_bytes = response.content

    # Open the PDF file
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))

    # Initialize variables to store the extracted row
    found_row = None

    # Loop through each page in the PDF
    for page in pdf.pages:
        text = page.extract_text()

        # Check if the search text exists in the current page
        if search_text in text:
            # Split the text into lines
            lines = text.split('\n')

            # Find the line index containing the search text
            for idx, line in enumerate(lines):
                if search_text in line:
                    # Extract the entire row
                    found_row = ' '.join(lines[idx:idx+4]).strip()  # Adjusted to include next 3 lines
                    break
            if found_row:
                break  # Exit loop if the first occurrence is found

    pdf.close()
    return found_row

# Function to create clickable links in Excel
def create_hyperlink(url, title):
    return f'=HYPERLINK("{url}", "{title}")'

# Function to convert extracted row to DataFrame
def row_to_dataframe(row, title, url, year):
    if row is None:
        return None

    # Split the row into columns based on whitespace
    columns = row.split()  # Change from row[0].split() to row.split()

    # Verify if the number of columns matches the expected number
    if len(columns) % 4 != 0:
        print(f"Invalid row data: {row}")
        return None

    # Group the columns into chunks of 4 to represent each row of data
    chunked_columns = [columns[i:i+4] for i in range(0, len(columns), 4)]

    # Create DataFrame with specified column names
    df = pd.DataFrame(chunked_columns, columns=["RE Generator", "Schedule (MU)", "Actual (MU)", "Deviation (MU)"])

    # Add year and PDF URL to DataFrame
    df["Year"] = year
    df["PDF URL"] = create_hyperlink(url, title)

    return df


# Function to perform search on multiple PDF URLs and append results into one DataFrame
def search_text_in_multiple_pdfs(pdf_links, search_text, year):
    all_rows = []
    for title, url in pdf_links:
        print(f"Searching {search_text} in {url}")

        found_row = search_text_in_pdf(title, url, search_text)
        if found_row:
            print("\t>> Found", found_row)
            all_rows.append((found_row, title, url))

    df = pd.concat([row_to_dataframe(row, title, url, year) for row, title, url in all_rows], ignore_index=True)
    return df

# Define a function to extract data based on selected year and title filter
def extract_data(year, title_filter):
    wrpc_base_url = "https://www.wrpc.gov.in"
    REA_link = f"{wrpc_base_url}/assets/data/REA_{year}.txt"
    response = requests.get(REA_link)
    if response.status_code == 200:
        rea_data = response.text.split("\n")
        pdf_links = []
        for data_line in rea_data:
            if ".pdf" in data_line:
                data_parts = data_line.split(",")
                if len(data_parts) >= 3:
                    month_year = data_parts[0].strip()
                    title = data_parts[0].strip() + ", " + data_parts[1].strip()
                    print("title", title)
                    print("title_filter.lower()", title_filter.lower())

                    link = wrpc_base_url + "/" + data_parts[2].strip()
                    if title_filter.lower() in title.lower():
                        print("matched")
                        pdf_links.append((title, link))
        for index, (title, link) in enumerate(pdf_links):
            print(f"{index}, \"{title}\", \"{link}\"")
        
        # Search for the text in the multiple PDFs and append the results into one DataFrame
        df = search_text_in_multiple_pdfs(pdf_links, "Arinsun_RUMS", year)  # You can adjust the search text as needed
        st.write(df)
        sheet_name = 'WRPC_Monthly Scheduled Revenue'
        create_file(df, sheet_name)

if __name__ == '__main__':
    # REGIONAL ENERGY ACCOUNTS
    st.markdown('### REGIONAL ENERGY ACCOUNTS')

    # Define the years and title filters for dropdown options
    years = ['2023', '2022', '2021', '2020', '2019', '2018', '2017', '2016', '2015', '2014', '2013', '2012', '2011', '2010']

    # Create dropdown widgets for selecting the year and title filter
    selected_year = st.selectbox('Select a Year:', years)
    current_month = datetime.now().strftime('%B')
    selected_month = st.text_input('Enter a Month:', current_month)

    # Create a button widget for triggering data extraction
    if st.button('Extract Data'):
        extract_data(selected_year, selected_month)
        st.write(f"Data extracted for: {selected_month}, {selected_year}")
        print('Done')
